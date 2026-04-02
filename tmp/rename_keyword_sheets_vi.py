# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from pathlib import Path

path = Path(r'D:/anti/perfume-luxury-vn/docs/reports/keyword-master-final-2026-03-30.xlsx')
wb = load_workbook(path)
rename_map = {
    '01_Master_All_Keywords': '01_Tat_ca_tu_khoa',
    '02_MAP_OK': '02_Map_dung_trang',
    '03_NEED_NEW_CONTENT': '03_Can_viet_bai_moi',
    '04_NEED_NEW_PRODUCT_PAGE': '04_Can_them_trang_san_pham',
    '05_UNMAPPED_REVIEW': '05_Chua_map_can_review',
    '06_Article_Keyword_Pool': '06_Pool_tu_khoa_bai_viet',
    '07_Extra_Diverse_Keywords': '07_Tu_khoa_mo_rong',
    '08_Summary': '08_Tong_quan',
}
for old, new in rename_map.items():
    if old in wb.sheetnames:
        wb[old].title = new
wb.save(path)
print(path)
