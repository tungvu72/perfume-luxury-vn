# -*- coding: utf-8 -*-
import pandas as pd
import re, unicodedata, pathlib, glob
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

workspace = pathlib.Path(r'D:/anti/perfume-luxury-vn')
out_path = workspace / 'docs' / 'reports' / 'keyword-master-final-2026-03-30.xlsx'
summary_path = workspace / 'docs' / 'reports' / 'keyword-master-final-2026-03-30-summary.md'


def norm(s: str) -> str:
    s = str(s).lower().strip()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = s.replace('đ', 'd')
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def comp_bucket(label, idx):
    if label == 'Thấp' or idx <= 35:
        return 'Thấp'
    if label == 'Trung bình' or idx <= 70:
        return 'Trung bình'
    return 'Cao'

src = glob.glob(r'd:/anti/ai-team/openclaw/state/coder/media/inbound/*.xlsx')[0]
df = pd.read_excel(src, header=2)
df = df.dropna(subset=['Keyword']).copy()
df['Keyword'] = df['Keyword'].astype(str).str.strip()
df['KeywordNorm'] = df['Keyword'].map(norm)
df['Volume'] = pd.to_numeric(df['Avg. monthly searches'], errors='coerce').fillna(0).astype(int)
df['CompetitionIndex'] = pd.to_numeric(df['Competition (indexed value)'], errors='coerce').fillna(999).astype(int)
df['Competition'] = [comp_bucket(str(l), i) for l, i in zip(df['Competition'].fillna('Không xác định'), df['CompetitionIndex'])]
df['WordCount'] = df['KeywordNorm'].str.split().str.len()

text = (workspace / 'src/constants/mockData.ts').read_text(encoding='utf-8')
products = []
for m in re.finditer(r'id:\s*"([^"]+)"\s*,\s*brandSlug:\s*"([^"]+)"\s*,\s*brand:\s*"([^"]+)"\s*,\s*name:\s*"([^"]+)"\s*,.*?gender:\s*"([^"]+)"', text, re.S):
    pid,bslug,brand,name,gender = m.groups()
    products.append({'id':pid,'brand':brand,'name':name,'brandNorm':norm(brand),'nameNorm':norm(name),'combo':norm(f'{brand} {name}'),'url':f'/nuoc-hoa-{gender}-{bslug}-{pid}'})
brand_pages = {'armani':'/armani','giorgio armani':'/armani','acqua di gio':'/armani','byredo':'/byredo','chanel':'/chanel','creed':'/creed','dior':'/dior','gucci':'/gucci','le labo':'/le-labo','prada':'/prada','versace':'/versace','ysl':'/ysl','yves saint laurent':'/ysl'}
articles = {
    'edt edp':'/edt-edp-la-gi',
    'nuoc hoa unisex la gi':'/nuoc-hoa-unisex-la-gi',
    'xit nuoc hoa dung cach':'/xit-nuoc-hoa-dung-cach-cho-nam',
    'nuoc hoa nam duoi 1 trieu':'/nuoc-hoa-nam-duoi-1-trieu-dang-mua-2026',
    'nuoc hoa nam di lam mua he':'/nuoc-hoa-nam-di-lam-mua-he',
    'nuoc hoa nam luu huong lau nhat mua he':'/nuoc-hoa-nam-luu-huong-lau-nhat-mua-he-2026',
    'top 7 nuoc hoa nam van phong lich lam':'/top-7-nuoc-hoa-nam-van-phong-lich-lam-2026',
    'lan dau mua nuoc hoa nam':'/lan-dau-mua-nuoc-hoa-nam-nen-chon-gi',
    'bleu de chanel edp vs ysl y edp':'/bleu-de-chanel-edp-vs-ysl-y-edp',
}

brand_aliases = sorted(set(list(brand_pages.keys()) + [p['brandNorm'] for p in products]), key=len, reverse=True)
product_aliases = sorted([(p['combo'], p) for p in products] + [(p['nameNorm'], p) for p in products], key=lambda x: len(x[0]), reverse=True)


def classify(kwn):
    if ' vs ' in f' {kwn} ' or ' so sanh ' in f' {kwn} ':
        return 'Article', 'Comparison', 'So sánh'
    if any(x in kwn for x in ['la gi','cach ','cach xit','huong dan','phan biet','that gia','fake','nam hay nu','cho nam hay nu']):
        return 'Article', 'Educational', 'Giải thích / how-to / trust'
    if any(x in kwn for x in ['review','danh gia']):
        return 'Article', 'Review', 'Review / đánh giá'
    if any(x in kwn for x in ['top ','mua he','mua dong','di lam','van phong','hen ho','date','luu huong','thom lau','duoi 1 trieu','duoi 2 trieu','cao cap','mini','10ml','20ml','chiet']):
        return 'Mixed', 'Need-based', 'Nhu cầu / roundup / mini'
    if any(x in kwn for x in ['chinh hang','authentic','gia bao nhieu','bao nhieu tien']) or any(b in kwn for b in brand_aliases):
        return 'Commercial', 'Transactional', 'Mua hàng / brand / product'
    if 'nuoc hoa nam' in kwn or 'nuoc hoa nu' in kwn or 'unisex' in kwn:
        return 'Commercial', 'Category', 'Danh mục chung'
    return 'Review', 'Unclear', 'Chưa rõ'

rows=[]
for _, r in df.iterrows():
    kw = r['Keyword']
    kwn = r['KeywordNorm']
    intent_group, sub_type, cluster = classify(kwn)
    target = ''
    status = 'UNMAPPED_REVIEW'
    note = ''

    # existing article exactish
    for token, url in articles.items():
        if token in kwn:
            target = url
            status = 'MAP_OK'
            note = 'Match bài viết hiện có'
            break

    # product
    if not target:
        for alias, p in product_aliases:
            if len(alias) >= 4 and alias in kwn:
                if intent_group in ['Commercial', 'Mixed', 'Article']:
                    target = p['url']
                    status = 'MAP_OK'
                    note = f"Match sản phẩm hiện có: {p['brand']} {p['name']}"
                break

    # brand/category mapping
    if not target:
        if 'nuoc hoa nu' in kwn and any(x in kwn for x in ['chinh hang','authentic']):
            target = '/nuoc-hoa-nu'
            status = 'MAP_OK'
            note = 'Keyword transactional nữ -> danh mục nữ'
        elif 'nuoc hoa nam' in kwn and any(x in kwn for x in ['chinh hang','authentic']):
            target = '/nuoc-hoa-nam'
            status = 'MAP_OK'
            note = 'Keyword transactional nam -> danh mục nam'
        elif 'nuoc hoa nam' in kwn and sub_type == 'Category':
            target = '/nuoc-hoa-nam'; status='MAP_OK'; note='Keyword generic danh mục nam'
        elif 'nuoc hoa nu' in kwn and sub_type == 'Category':
            target = '/nuoc-hoa-nu'; status='MAP_OK'; note='Keyword generic danh mục nữ'
        elif 'unisex' in kwn and sub_type == 'Category':
            target = '/unisex'; status='MAP_OK'; note='Keyword generic danh mục unisex'
        else:
            for b, url in brand_pages.items():
                if re.search(rf'(^| ){re.escape(b)}( |$)', kwn):
                    if any(x in kwn for x in ['chinh hang','authentic','gia','review','danh gia']) or sub_type == 'Transactional':
                        target = url; status='MAP_OK'; note=f'Match brand page: {b}'
                        break

    if not target:
        if intent_group == 'Article':
            status = 'NEED_NEW_CONTENT'
            note = 'Intent bài viết rõ nhưng chưa có landing page phù hợp'
        elif intent_group == 'Mixed':
            # split mixed: if transactional modifiers then category/product else content
            if any(x in kwn for x in ['mini','10ml','20ml']):
                status = 'NEED_NEW_CONTENT'
                note = 'Cluster mini/chiết cần bài hoặc hub riêng'
            elif any(x in kwn for x in ['mua he','di lam','van phong','hen ho','date','luu huong','duoi 1 trieu','duoi 2 trieu','cao cap']):
                status = 'NEED_NEW_CONTENT'
                note = 'Cluster nhu cầu nên làm bài roundup/hub riêng'
            else:
                status = 'UNMAPPED_REVIEW'
        elif intent_group == 'Commercial':
            status = 'NEED_NEW_PRODUCT_PAGE'
            note = 'Keyword transactional nhưng chưa có page đích chuẩn trên web'
        else:
            status = 'UNMAPPED_REVIEW'
            note = 'Chưa đủ chắc để map an toàn'

    rows.append({
        'Keyword': kw,
        'Volume': r['Volume'],
        'Competition': r['Competition'],
        'CompetitionIndex': r['CompetitionIndex'],
        'IntentGroup': intent_group,
        'KeywordType': sub_type,
        'Cluster': cluster,
        'Status': status,
        'TargetPage': target,
        'Note': note,
    })

master_df = pd.DataFrame(rows)
map_ok_df = master_df[master_df['Status']=='MAP_OK'].sort_values(['Volume','CompetitionIndex'], ascending=[False,True])
content_df = master_df[master_df['Status']=='NEED_NEW_CONTENT'].sort_values(['Volume','CompetitionIndex'], ascending=[False,True])
product_df = master_df[master_df['Status']=='NEED_NEW_PRODUCT_PAGE'].sort_values(['Volume','CompetitionIndex'], ascending=[False,True])
unmapped_df = master_df[master_df['Status']=='UNMAPPED_REVIEW'].sort_values(['Volume','CompetitionIndex'], ascending=[False,True])

extra_keywords = [
    ['nước hoa mùi sạch cho nam','Article','Buying guide','Từ khóa lạ / clean scent'],
    ['nước hoa mùi bánh ngọt cho nữ','Article','Buying guide','Theo mùi hương cụ thể'],
    ['nước hoa cho da dầu có bám lâu không','Article','Educational','Problem-solving'],
    ['xịt nước hoa lên áo hay da','Article','Educational','FAQ evergreen'],
    ['nên mua chiết hay fullbox nước hoa','Article','Educational','FAQ chuyển đổi cao'],
    ['nước hoa mùa mưa sài gòn','Article','Buying guide','Local context'],
    ['nước hoa nam đi cafe ban ngày','Article','Buying guide','Long-tail lạ'],
    ['nước hoa nam mùi trà','Article','Buying guide','Theo note mùi'],
    ['nước hoa nữ mùi xà phòng','Article','Buying guide','Theo vibe mùi'],
    ['nước hoa niche cho người mới','Article','Educational','Niche onboarding'],
    ['nước hoa đi biển cho nam','Article','Buying guide','Use case mùa hè'],
    ['nước hoa nữ văn phòng thơm nhẹ','Article','Buying guide','Cluster nữ thiếu'],
    ['nước hoa unisex dưới 2 triệu','Article','Buying guide','Budget + unisex'],
    ['nước hoa giống mùi nước xả vải','Article','Buying guide','Keyword lạ dễ kéo traffic'],
    ['nước hoa mùi sang như khách sạn 5 sao','Article','Buying guide','Keyword lạ theo vibe'],
    ['nước hoa cho người không thích mùi nồng','Article','Buying guide','Pain-point keyword'],
    ['nước hoa đi đám cưới cho nam','Article','Buying guide','Use case cụ thể'],
    ['nước hoa cho phòng máy lạnh','Article','Educational','Context-based keyword'],
    ['nước hoa mùi lạnh cho nam','Article','Buying guide','Style-based keyword'],
    ['nước hoa nữ mùi trà trắng','Article','Buying guide','Long-tail note based'],
]
extra_df = pd.DataFrame(extra_keywords, columns=['KeywordIdea','IntentGroup','KeywordType','WhyAdd'])

article_pool = pd.concat([
    content_df[['Keyword','Volume','Competition','CompetitionIndex','Cluster']].rename(columns={'Keyword':'PrimaryKeyword'}).head(120),
    extra_df.rename(columns={'KeywordIdea':'PrimaryKeyword'})[['PrimaryKeyword','IntentGroup','KeywordType','WhyAdd']]
], ignore_index=True)

summary_df = master_df.groupby(['Status','IntentGroup','KeywordType'], dropna=False).agg(KeywordCount=('Keyword','count'), TotalVolume=('Volume','sum')).reset_index().sort_values(['Status','TotalVolume'], ascending=[True,False])

with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    master_df.to_excel(writer, sheet_name='01_Master_All_Keywords', index=False)
    map_ok_df.to_excel(writer, sheet_name='02_MAP_OK', index=False)
    content_df.to_excel(writer, sheet_name='03_NEED_NEW_CONTENT', index=False)
    product_df.to_excel(writer, sheet_name='04_NEED_NEW_PRODUCT_PAGE', index=False)
    unmapped_df.to_excel(writer, sheet_name='05_UNMAPPED_REVIEW', index=False)
    article_pool.to_excel(writer, sheet_name='06_Article_Keyword_Pool', index=False)
    extra_df.to_excel(writer, sheet_name='07_Extra_Diverse_Keywords', index=False)
    summary_df.to_excel(writer, sheet_name='08_Summary', index=False)

wb = load_workbook(out_path)
fill = PatternFill('solid', fgColor='1F4E78')
white = Font(color='FFFFFF', bold=True, name='Calibri')
body = Font(name='Calibri')
side = Side(style='thin', color='D9E2F3')
for ws in wb.worksheets:
    ws.freeze_panes = 'A2'
    for c in ws[1]:
        c.fill = fill; c.font = white; c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); c.border = Border(top=side,bottom=side,left=side,right=side)
    widths = {}
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = body; c.alignment = Alignment(vertical='top', wrap_text=True); c.border = Border(top=side,bottom=side,left=side,right=side)
            val = '' if c.value is None else str(c.value)
            widths[c.column_letter] = min(max(widths.get(c.column_letter, 10), len(val)+2), 42)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
wb.save(out_path)

summary = f'''# Keyword master final\n\n- Tổng keyword planner: **{len(master_df)}**\n- MAP_OK: **{len(map_ok_df)}**\n- NEED_NEW_CONTENT: **{len(content_df)}**\n- NEED_NEW_PRODUCT_PAGE: **{len(product_df)}**\n- UNMAPPED_REVIEW: **{len(unmapped_df)}**\n- Extra diverse keywords added: **{len(extra_df)}**\n\n## Ghi chú\n- Đã sửa rule intent: keyword transactional như "nước hoa nữ chính hãng" map về category, không đẩy sang article.\n- Sheet `06_Article_Keyword_Pool` = pool để viết bài, gồm keyword planner phù hợp + keyword mở rộng lạ hơn.\n- Mày có thể gửi thêm batch keyword sau để tao append vào file này vòng 2.\n'''
summary_path.write_text(summary, encoding='utf-8')
print(out_path)
