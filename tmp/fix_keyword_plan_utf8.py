# -*- coding: utf-8 -*-
import pandas as pd
import re
import unicodedata
import pathlib
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

workspace = pathlib.Path(r'D:/anti/perfume-luxury-vn')
out_path = workspace / 'docs' / 'reports' / 'keyword-plan-2026-03-30-v3.xlsx'
summary_path = workspace / 'docs' / 'reports' / 'keyword-plan-2026-03-30-v3-summary.md'


def norm(s: str) -> str:
    s = str(s).lower().strip()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = s.replace('đ', 'd')
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


input_files = glob.glob(r'd:/anti/ai-team/openclaw/state/coder/media/inbound/*.xlsx')
input_path = input_files[0]

df = pd.read_excel(input_path, header=2)
df = df.dropna(subset=['Keyword']).copy()
df['Keyword'] = df['Keyword'].astype(str).str.strip()
df['KeywordNorm'] = df['Keyword'].map(norm)
df['AvgMonthlySearches'] = pd.to_numeric(df['Avg. monthly searches'], errors='coerce').fillna(0).astype(int)
df['CompetitionIndex'] = pd.to_numeric(df['Competition (indexed value)'], errors='coerce').fillna(999).astype(int)
df['CompetitionLabel'] = df['Competition'].fillna('Không xác định').astype(str)
df['WordCount'] = df['KeywordNorm'].str.split().str.len()


def comp_bucket(label, idx):
    if label == 'Thấp' or idx <= 35:
        return 'Thấp'
    if label == 'Trung bình' or idx <= 70:
        return 'Trung bình'
    return 'Cao'


df['CompetitionBucket'] = [comp_bucket(l, i) for l, i in zip(df['CompetitionLabel'], df['CompetitionIndex'])]

text = (workspace / 'src/constants/mockData.ts').read_text(encoding='utf-8')
products = []
pattern = re.finditer(r'id:\s*"([^"]+)"\s*,\s*brandSlug:\s*"([^"]+)"\s*,\s*brand:\s*"([^"]+)"\s*,\s*name:\s*"([^"]+)"\s*,.*?gender:\s*"([^"]+)"', text, re.S)
for m in pattern:
    pid, bslug, brand, name, gender = m.groups()
    products.append({
        'id': pid,
        'brandSlug': bslug,
        'brand': brand,
        'name': name,
        'gender': gender,
        'url': f'/nuoc-hoa-{gender}-{bslug}-{pid}',
        'nameNorm': norm(name),
        'brandNorm': norm(brand),
        'comboNorm': norm(f'{brand} {name}'),
        'fullNorm': norm(f'nuoc hoa {brand} {name}')
    })

brand_pages = {
    'armani': '/armani',
    'giorgio armani': '/armani',
    'acqua di gio': '/armani',
    'byredo': '/byredo',
    'chanel': '/chanel',
    'creed': '/creed',
    'dior': '/dior',
    'gucci': '/gucci',
    'le labo': '/le-labo',
    'prada': '/prada',
    'versace': '/versace',
    'ysl': '/ysl',
    'yves saint laurent': '/ysl'
}

existing_articles = [
    ('bleu de chanel edp vs ysl y edp', '/bleu-de-chanel-edp-vs-ysl-y-edp', 'Bài so sánh hiện có'),
    ('cach chon nuoc hoa nam di lam trong khi hau nong am viet nam', '/cach-chon-nuoc-hoa-nam-di-lam-trong-khi-hau-nong-am-viet-nam', 'Bài kiến thức hiện có'),
    ('creed la thuong hieu nuoc hoa nhu the nao', '/creed-la-thuong-hieu-nuoc-hoa-nhu-the-nao', 'Bài brand story hiện có'),
    ('edt edp', '/edt-edp-la-gi', 'Bài kiến thức hiện có'),
    ('lan dau mua nuoc hoa nam', '/lan-dau-mua-nuoc-hoa-nam-nen-chon-gi', 'Bài kiến thức hiện có'),
    ('nuoc hoa nam di lam mua he', '/nuoc-hoa-nam-di-lam-mua-he', 'Bài roundup hiện có'),
    ('nuoc hoa nam duoi 1 trieu', '/nuoc-hoa-nam-duoi-1-trieu-dang-mua-2026', 'Bài roundup hiện có'),
    ('nuoc hoa nam luu huong lau nhat mua he', '/nuoc-hoa-nam-luu-huong-lau-nhat-mua-he-2026', 'Bài roundup hiện có'),
    ('nuoc hoa unisex la gi', '/nuoc-hoa-unisex-la-gi', 'Bài kiến thức hiện có'),
    ('top 7 nuoc hoa nam di lam mua he', '/top-7-nuoc-hoa-nam-di-lam-mua-he-2026', 'Bài roundup hiện có'),
    ('top 7 nuoc hoa nam van phong lich lam', '/top-7-nuoc-hoa-nam-van-phong-lich-lam-2026', 'Bài roundup hiện có'),
    ('xit nuoc hoa dung cach cho nam', '/xit-nuoc-hoa-dung-cach-cho-nam', 'Bài kiến thức hiện có'),
]

mapped = []
for _, r in df.iterrows():
    kw = r['Keyword']
    kwn = r['KeywordNorm']
    page_type = ''
    page_url = ''
    page_name = ''
    intent_group = ''
    action = ''
    reason = ''
    priority = 4

    for token, url, ptype in existing_articles:
        if token in kwn:
            page_type = ptype
            page_url = url
            page_name = url.strip('/')
            intent_group = 'Informational'
            action = 'Giữ và tối ưu internal link'
            reason = f'Khớp intent với bài đang có: {token}'
            priority = 1
            break

    if not page_url:
        best = None
        for p in products:
            score = 0
            if kwn == p['fullNorm'] or kwn == p['comboNorm']:
                score = 100
            elif p['fullNorm'] in kwn or p['comboNorm'] in kwn:
                score = 90
            elif kwn == p['nameNorm'] or (p['nameNorm'] in kwn and len(p['nameNorm'].split()) >= 2):
                score = 80
            if score and (best is None or score > best[0]):
                best = (score, p)
        if best:
            p = best[1]
            mod = ''
            if any(x in kwn for x in ['review', 'danh gia']):
                mod = ' + nên có bài review hỗ trợ'
            elif any(x in kwn for x in ['chinh hang', 'authentic', 'that gia', 'fake', 'mini', '10ml', '20ml', 'edp', 'edt', 'parfum', 'gia']):
                mod = ' + nên có content bổ trợ theo intent'
            page_type = 'Product page hiện có'
            page_url = p['url']
            page_name = p['name']
            intent_group = 'Commercial / Product'
            action = 'Map về product page' + mod
            reason = f"Khớp sản phẩm: {p['brand']} {p['name']}"
            priority = 1

    if not page_url:
        for b, url in sorted(brand_pages.items(), key=lambda x: len(x[0]), reverse=True):
            if re.search(rf'(^| ){re.escape(b)}( |$)', kwn):
                page_type = 'Brand page hiện có'
                page_url = url
                page_name = url.strip('/')
                intent_group = 'Brand discovery'
                action = 'Map về brand page'
                reason = f'Khớp brand: {b}'
                priority = 2
                break

    if not page_url:
        if any(x in kwn for x in ['nuoc hoa nam', 'nuoc hoa gio nam']) and not any(x in kwn for x in ['mini', '10ml', '20ml', 'chinh hang', 'authentic', 'that gia', 'fake', 'review', 'danh gia', 'mua he', 'van phong', 'di lam']):
            page_type = 'Listing hiện có'
            page_url = '/nuoc-hoa-nam'
            page_name = 'Nước hoa nam'
            intent_group = 'Generic category'
            action = 'Map về listing nam'
            reason = 'Keyword generic nhóm nam'
            priority = 2
        elif any(x in kwn for x in ['nuoc hoa nu', 'nuoc hoa nua']) and not any(x in kwn for x in ['mini', '10ml', '20ml', 'chinh hang', 'authentic', 'that gia', 'fake', 'review', 'danh gia', 'mua he', 'van phong', 'di lam']):
            page_type = 'Listing hiện có'
            page_url = '/nuoc-hoa-nu'
            page_name = 'Nước hoa nữ'
            intent_group = 'Generic category'
            action = 'Map về listing nữ'
            reason = 'Keyword generic nhóm nữ'
            priority = 2
        elif 'unisex' in kwn and not any(x in kwn for x in ['mua he', 'mini', 'chinh hang', 'review']):
            page_type = 'Listing hiện có'
            page_url = '/unisex'
            page_name = 'Nước hoa unisex'
            intent_group = 'Generic category'
            action = 'Map về listing unisex'
            reason = 'Keyword generic nhóm unisex'
            priority = 2
        elif any(x in kwn for x in ['mua he', 'van phong', 'di lam', 'hen ho', 'date', 'luu huong', 'thom lau', 'cao cap', 'duoi 1 trieu', 'duoi 2 trieu']):
            page_type = 'Hub nhu cầu hiện có'
            page_url = '/nuoc-hoa-theo-nhu-cau'
            page_name = 'Nước hoa theo nhu cầu'
            intent_group = 'Need-based discovery'
            action = 'Map về hub nhu cầu + nên viết bài riêng nếu volume tốt'
            reason = 'Keyword theo hoàn cảnh hoặc nhu cầu sử dụng'
            priority = 3
        elif any(x in kwn for x in ['edt', 'edp', 'parfum', 'xit nuoc hoa', 'la gi', 'cho nam hay nu', 'that gia', 'authentic', 'mini', '10ml', '20ml', 'chinh hang']):
            page_type = 'Cần bài viết riêng'
            page_url = ''
            page_name = ''
            intent_group = 'Informational / long-tail'
            action = 'Ưu tiên content mới'
            reason = 'Intent dài, nên SEO bằng bài viết chuyên biệt'
            priority = 1
        else:
            page_type = 'Trang chủ / hub tổng hợp'
            page_url = '/'
            page_name = 'Trang chủ'
            intent_group = 'Broad discovery'
            action = 'Chưa nên ưu tiên riêng'
            reason = 'Keyword quá rộng hoặc chưa đủ rõ intent'
            priority = 4

    mapped.append({
        'Keyword': kw,
        'AvgMonthlySearches': r['AvgMonthlySearches'],
        'Competition': r['CompetitionBucket'],
        'CompetitionIndex': r['CompetitionIndex'],
        'WordCount': r['WordCount'],
        'IntentGroup': intent_group,
        'MappedPageType': page_type,
        'MappedPageURL': page_url,
        'MappedPageName': page_name,
        'RecommendedAction': action,
        'Reason': reason,
        'Priority': priority,
    })

mapping_df = pd.DataFrame(mapped).sort_values(
    ['Priority', 'AvgMonthlySearches', 'CompetitionIndex', 'WordCount'],
    ascending=[True, False, True, False]
)

summary_df = mapping_df.groupby(['MappedPageType', 'MappedPageURL', 'MappedPageName'], dropna=False).agg(
    KeywordCount=('Keyword', 'count'),
    TotalSearchVolume=('AvgMonthlySearches', 'sum'),
    AvgCompetitionIndex=('CompetitionIndex', 'mean')
).reset_index().sort_values(['TotalSearchVolume', 'KeywordCount'], ascending=[False, False])

content_plan = pd.DataFrame([
    ['nước hoa nữ chính hãng', 5400, 'Thấp', 5, 'nuoc-hoa-nu-chinh-hang-shop-uy-tin-2026', 'Nước hoa nữ chính hãng: cách chọn shop uy tín và tránh hàng fake 2026', 'Buying guide', 'Commercial + trust', 'nước hoa chính hãng nữ; nuoc hoa nam chinh hang; nước hoa dior nữ chính hãng; nước hoa ysl nữ chính hãng', 1],
    ['nước hoa cao cấp', 4400, 'Thấp', 9, 'nuoc-hoa-cao-cap-dang-tien-2026', 'Nước hoa cao cấp là gì? 12 chai đáng tiền theo từng gu 2026', 'Buying guide', 'Commercial investigation', 'nước hoa nữ cao cấp; nước hoa cao cấp nữ; nước hoa cao cấp nam; nước hoa cao cấp cho nữ', 2],
    ['nước hoa edt và edp', 110, 'Cao', 87, 'edt-edp-parfum-khac-nhau-gi-2026', 'EDT, EDP, Parfum khác nhau gì? Cách chọn đúng theo nhu cầu 2026', 'Educational', 'Informational', 'nước hoa chloe eau de parfum; nước hoa chanel chance edp; nước hoa dior parfum', 3],
    ['nước hoa mini chính hãng', 110, 'Cao', 100, 'nuoc-hoa-mini-chinh-hang-dang-mua-2026', 'Nước hoa mini chính hãng: các line đáng mua nhất 2026', 'Buying guide', 'Commercial investigation', 'nước hoa mini 5ml chính hãng; nước hoa ysl mini; nước hoa chanel mini; nước hoa lv mini', 4],
    ['nước hoa nam mini', 30, 'Cao', 97, 'nuoc-hoa-nam-mini-dang-mua-2026', 'Nước hoa nam mini đáng mua: 10 lựa chọn test mùi trước khi mua fullsize 2026', 'Buying guide', 'Commercial investigation', 'nước hoa nam 10ml; nước hoa dior sauvage 10ml', 5],
    ['nước hoa nam authentic', 30, 'Cao', 94, 'nuoc-hoa-authentic-la-gi-2026', 'Nước hoa authentic là gì? Cách hiểu đúng để không mua hớ 2026', 'Educational', 'Trust + informational', 'authentic nước hoa; nước hoa authentic nữ', 6],
    ['nước hoa unisex mùa hè', 20, 'Cao', 100, 'top-nuoc-hoa-unisex-de-dung-2026', 'Top nước hoa unisex dễ dùng nhất cho khí hậu Việt Nam 2026', 'Buying guide', 'Commercial investigation', 'nuoc hoa unisex; nước hoa unisex cho nữ', 7],
    ['phan biet nuoc hoa chanel that gia', 10, 'Thấp', 14, 'phan-biet-nuoc-hoa-that-gia-2026', 'Cách phân biệt nước hoa thật giả: 9 dấu hiệu người mới dễ bỏ qua 2026', 'Educational', 'Trust + informational', 'phân biệt versace thật giả; phân biệt nước hoa narciso thật giả; phân biệt chanel chance thật giả', 8],
    ['dior sauvage của nam hay nữ', 10, 'Thấp', 0, 'nuoc-hoa-nam-hay-nu-cach-doc-2026', 'Nước hoa dành cho nam hay nữ? Cách đọc nhanh theo từng chai nổi tiếng 2026', 'Educational', 'Informational', 'bleu de chanel của nam hay nữ; le labo 13 nam hay nữ; santal 33 cho nam hay nữ', 9],
], columns=['PrimaryKeyword', 'PrimaryVolume', 'Competition', 'CompetitionIndex', 'SuggestedSlug', 'ArticleTitleSuggestion', 'PageType', 'SearchIntent', 'SecondaryKeywords', 'Priority'])

extra_df = pd.DataFrame([
    ['nước hoa nữ văn phòng thơm nhẹ', 'Commercial investigation', 'Gap lớn cho nhóm nữ đi làm', 'Bài roundup mới'],
    ['nước hoa nữ mùa hè lưu hương lâu', 'Commercial investigation', 'Mirror cluster nam đang có', 'Bài roundup mới'],
    ['nước hoa nữ dưới 2 triệu', 'Commercial investigation', 'Budget cluster nữ còn trống', 'Bài roundup mới'],
    ['nước hoa đi hẹn hò cho nam', 'Commercial investigation', 'Intent chuyển đổi tốt', 'Bài roundup mới'],
    ['nước hoa mùi sạch cho nam', 'Commercial investigation', 'Liên quan office / clean scent', 'Bài roundup mới'],
    ['nên mua chiết hay fullbox nước hoa', 'Informational', 'FAQ sát chuyển đổi', 'Bài educational'],
    ['xịt nước hoa lên áo hay da', 'Informational', 'FAQ evergreen', 'Bài educational'],
    ['nước hoa niche cho người mới', 'Informational', 'Bổ trợ cluster niche', 'Bài guide mới'],
    ['nước hoa mùa mưa sài gòn', 'Commercial investigation', 'Local context mạnh', 'Bài roundup mới'],
    ['nước hoa cho da dầu có bám lâu không', 'Informational', 'Problem-solving content', 'Bài educational'],
], columns=['KeywordIdea', 'Intent', 'WhyAdd', 'RecommendedPage'])

top100_df = mapping_df.head(100).copy()

with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    mapping_df.to_excel(writer, sheet_name='Keyword Mapping', index=False)
    summary_df.to_excel(writer, sheet_name='Page Summary', index=False)
    content_plan.to_excel(writer, sheet_name='Daily Content Plan', index=False)
    extra_df.to_excel(writer, sheet_name='Extra Keyword Ideas', index=False)
    top100_df.to_excel(writer, sheet_name='Top 100 Priorities', index=False)

wb = load_workbook(out_path)
header_fill = PatternFill('solid', fgColor='1F4E78')
header_font = Font(color='FFFFFF', bold=True, name='Calibri')
body_font = Font(name='Calibri')
thin = Side(style='thin', color='D9E2F3')
for ws in wb.worksheets:
    ws.freeze_panes = 'A2'
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = body_font
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            val = '' if c.value is None else str(c.value)
            widths[c.column_letter] = min(max(widths.get(c.column_letter, 10), len(val) + 2), 42)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
wb.save(out_path)

summary = """# Keyword plan v3

- Source rows: **3865**
- Mapped rows: **3865**
- New content priorities: **9**

## Fixes
- Đã xuất lại file bằng script UTF-8 riêng
- Đã sửa lỗi text bị `n??c hoa` / `hi?n c?`
- Đã giữ cấu trúc sheet logic hơn: mapping, summary, daily content, extra ideas, top priorities
- Đã sort theo Priority -> Volume -> Competition

## Top content nên viết trước
- nước hoa nữ chính hãng (5400) -> /nuoc-hoa-nu-chinh-hang-shop-uy-tin-2026
- nước hoa cao cấp (4400) -> /nuoc-hoa-cao-cap-dang-tien-2026
- nước hoa edt và edp (110) -> /edt-edp-parfum-khac-nhau-gi-2026
- nước hoa mini chính hãng (110) -> /nuoc-hoa-mini-chinh-hang-dang-mua-2026
- nước hoa nam mini (30) -> /nuoc-hoa-nam-mini-dang-mua-2026
"""
summary_path.write_text(summary, encoding='utf-8')
print(out_path)
print(summary_path)
