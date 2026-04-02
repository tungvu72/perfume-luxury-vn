# -*- coding: utf-8 -*-
import pandas as pd
import re
import unicodedata
import pathlib
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

workspace = pathlib.Path(r'D:/anti/perfume-luxury-vn')
out_path = workspace / 'docs' / 'reports' / 'keyword-plan-2026-03-30-final.xlsx'
summary_path = workspace / 'docs' / 'reports' / 'keyword-plan-2026-03-30-final-summary.md'


def norm(s: str) -> str:
    s = str(s).lower().strip()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = s.replace('đ', 'd')
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def comp_bucket(label, idx):
    if label == 'Thấp' or idx <= 35:
        return 'Thấp'
    if label == 'Trung bình' or idx <= 70:
        return 'Trung bình'
    return 'Cao'


# ---------- Load source ----------
input_path = glob.glob(r'd:/anti/ai-team/openclaw/state/coder/media/inbound/*.xlsx')[0]
df = pd.read_excel(input_path, header=2)
df = df.dropna(subset=['Keyword']).copy()
df['Keyword'] = df['Keyword'].astype(str).str.strip()
df['KeywordNorm'] = df['Keyword'].map(norm)
df['Volume'] = pd.to_numeric(df['Avg. monthly searches'], errors='coerce').fillna(0).astype(int)
df['CompetitionIndex'] = pd.to_numeric(df['Competition (indexed value)'], errors='coerce').fillna(999).astype(int)
df['CompetitionRaw'] = df['Competition'].fillna('Không xác định').astype(str)
df['Competition'] = [comp_bucket(l, i) for l, i in zip(df['CompetitionRaw'], df['CompetitionIndex'])]
df['WordCount'] = df['KeywordNorm'].str.split().str.len()

# ---------- Load website inventory ----------
text = (workspace / 'src/constants/mockData.ts').read_text(encoding='utf-8')
products = []
for m in re.finditer(r'id:\s*"([^"]+)"\s*,\s*brandSlug:\s*"([^"]+)"\s*,\s*brand:\s*"([^"]+)"\s*,\s*name:\s*"([^"]+)"\s*,.*?gender:\s*"([^"]+)"', text, re.S):
    pid, bslug, brand, name, gender = m.groups()
    products.append({
        'id': pid,
        'brandSlug': bslug,
        'brand': brand,
        'name': name,
        'gender': gender,
        'url': f'/nuoc-hoa-{gender}-{bslug}-{pid}',
        'nameNorm': norm(name),
        'brandNorm': norm(brand),
        'comboNorm': norm(f'{brand} {name}'),
        'fullNorm': norm(f'nuoc hoa {brand} {name}')
    })

brand_pages = {
    'armani': '/armani',
    'giorgio armani': '/armani',
    'acqua di gio': '/armani',
    'byredo': '/byredo',
    'chanel': '/chanel',
    'creed': '/creed',
    'dior': '/dior',
    'gucci': '/gucci',
    'le labo': '/le-labo',
    'prada': '/prada',
    'versace': '/versace',
    'ysl': '/ysl',
    'yves saint laurent': '/ysl'
}

existing_articles = [
    ('bleu de chanel edp vs ysl y edp', '/bleu-de-chanel-edp-vs-ysl-y-edp', 'Comparison'),
    ('cach chon nuoc hoa nam di lam trong khi hau nong am viet nam', '/cach-chon-nuoc-hoa-nam-di-lam-trong-khi-hau-nong-am-viet-nam', 'Educational'),
    ('creed la thuong hieu nuoc hoa nhu the nao', '/creed-la-thuong-hieu-nuoc-hoa-nhu-the-nao', 'Brand story'),
    ('edt edp', '/edt-edp-la-gi', 'Educational'),
    ('lan dau mua nuoc hoa nam', '/lan-dau-mua-nuoc-hoa-nam-nen-chon-gi', 'Educational'),
    ('nuoc hoa nam di lam mua he', '/nuoc-hoa-nam-di-lam-mua-he', 'Buying guide'),
    ('nuoc hoa nam duoi 1 trieu', '/nuoc-hoa-nam-duoi-1-trieu-dang-mua-2026', 'Buying guide'),
    ('nuoc hoa nam luu huong lau nhat mua he', '/nuoc-hoa-nam-luu-huong-lau-nhat-mua-he-2026', 'Buying guide'),
    ('nuoc hoa unisex la gi', '/nuoc-hoa-unisex-la-gi', 'Educational'),
    ('top 7 nuoc hoa nam di lam mua he', '/top-7-nuoc-hoa-nam-di-lam-mua-he-2026', 'Buying guide'),
    ('top 7 nuoc hoa nam van phong lich lam', '/top-7-nuoc-hoa-nam-van-phong-lich-lam-2026', 'Buying guide'),
    ('xit nuoc hoa dung cach cho nam', '/xit-nuoc-hoa-dung-cach-cho-nam', 'Educational'),
]

brand_aliases = sorted(set(list(brand_pages.keys()) + [p['brandNorm'] for p in products]), key=len, reverse=True)
product_aliases = []
for p in products:
    product_aliases.extend([
        (p['fullNorm'], p),
        (p['comboNorm'], p),
        (p['nameNorm'], p),
    ])
product_aliases = sorted(product_aliases, key=lambda x: len(x[0]), reverse=True)
existing_product_names = {p['nameNorm'] for p in products}

# ---------- Rules ----------
def detect_intent_cluster(kwn: str):
    if ' vs ' in f' {kwn} ' or ' so sanh ' in f' {kwn} ':
        return 'Comparison', 'So sánh'
    if any(x in kwn for x in ['review', 'danh gia']):
        return 'Product review', 'Review / đánh giá'
    if any(x in kwn for x in ['that gia', 'fake', 'authentic', 'chinh hang']):
        return 'Trust', 'Chính hãng / authentic / thật giả'
    if any(x in kwn for x in ['edt', 'edp', 'parfum', 'eau de parfum']):
        return 'Informational', 'EDT / EDP / Parfum'
    if any(x in kwn for x in ['mini', '10ml', '20ml', 'chiet']):
        return 'Commercial investigation', 'Mini / chiết / dung tích nhỏ'
    if any(x in kwn for x in ['mua he', 'mua dong', 'di lam', 'van phong', 'hen ho', 'date', 'luu huong', 'thom lau', 'duoi 1 trieu', 'duoi 2 trieu', 'cao cap']):
        return 'Problem / need', 'Nhu cầu sử dụng'
    if any(x in kwn for x in ['la gi', 'xit nuoc hoa', 'cach xit', 'cach dung', 'nam hay nu', 'cho nam hay nu']):
        return 'Informational', 'Giải thích / how-to'
    if any(x in kwn for x in brand_aliases):
        return 'Brand / product commercial', 'Brand / product'
    if 'nuoc hoa nam' in kwn or 'nuoc hoa nu' in kwn or 'unisex' in kwn:
        return 'Generic category', 'Danh mục chung'
    return 'Unclear', 'Chưa rõ'


def find_existing_article(kwn: str):
    for token, url, typ in existing_articles:
        if token in kwn:
            return {'url': url, 'type': typ, 'token': token}
    return None


def find_existing_product(kwn: str):
    for alias, p in product_aliases:
        if len(alias) < 4:
            continue
        if kwn == alias or alias in kwn:
            return p
    return None


def detect_missing_product(kwn: str):
    # has product-commercial modifiers but no existing product match
    product_mods = ['review', 'danh gia', 'gia', 'chinh hang', 'authentic', 'that gia', 'fake', 'mini', '10ml', '20ml', 'edp', 'edt', 'parfum']
    if not any(x in kwn for x in product_mods):
        return None
    for b in brand_aliases:
        if re.search(rf'(^| ){re.escape(b)}( |$)', kwn):
            # remove brand token and generic modifiers to estimate product phrase
            stripped = kwn
            for tok in [b, 'nuoc hoa', 'review', 'danh gia', 'gia', 'bao nhieu', 'chinh hang', 'authentic', 'that gia', 'fake', 'mini', '10ml', '20ml', 'edp', 'edt', 'parfum', 'eau de parfum', 'cho nam hay nu', 'nam hay nu']:
                stripped = stripped.replace(tok, ' ')
            stripped = re.sub(r'\s+', ' ', stripped).strip()
            if len(stripped) >= 3 and stripped not in existing_product_names:
                return {'brand': b, 'candidate': stripped}
    return None


def priority_of(status, volume, comp_index):
    if status == 'MAP_OK':
        base = 1
    elif status == 'NEED_NEW_CONTENT':
        base = 2
    elif status == 'NEED_NEW_PRODUCT_PAGE':
        base = 3
    else:
        base = 4
    bonus = 0
    if volume >= 1000:
        bonus = -0.3
    elif volume >= 100:
        bonus = -0.2
    elif volume >= 10:
        bonus = -0.1
    comp_adj = comp_index / 1000.0
    return round(base + bonus + comp_adj, 3)


# ---------- Build strict mapping ----------
rows = []
for _, r in df.iterrows():
    kw = r['Keyword']
    kwn = r['KeywordNorm']
    intent_type, cluster = detect_intent_cluster(kwn)

    status = 'UNMAPPED_REVIEW'
    target_page = ''
    target_type = ''
    why = ''
    note = ''

    article = find_existing_article(kwn)
    product = find_existing_product(kwn)
    missing_product = detect_missing_product(kwn) if product is None else None

    if article:
        status = 'MAP_OK'
        target_page = article['url']
        target_type = f'Article ({article["type"]})'
        why = f'Khớp đúng intent với bài hiện có: {article["token"]}'
    elif product and intent_type in ['Product review', 'Brand / product commercial', 'Commercial investigation', 'Trust']:
        status = 'MAP_OK'
        target_page = product['url']
        target_type = 'Product page'
        why = f'Khớp đúng sản phẩm hiện có: {product["brand"]} {product["name"]}'
        if any(x in kwn for x in ['review', 'danh gia', 'that gia', 'fake', 'authentic', 'chinh hang', 'mini', '10ml', '20ml', 'edt', 'edp', 'parfum']):
            note = 'Có thể cần thêm content bổ trợ theo intent riêng, nhưng product page hiện vẫn là landing gần nhất.'
    elif intent_type == 'Generic category':
        if 'nuoc hoa nam' in kwn:
            status = 'MAP_OK'
            target_page = '/nuoc-hoa-nam'
            target_type = 'Category page'
            why = 'Keyword generic danh mục nam'
        elif 'nuoc hoa nu' in kwn:
            status = 'MAP_OK'
            target_page = '/nuoc-hoa-nu'
            target_type = 'Category page'
            why = 'Keyword generic danh mục nữ'
        elif 'unisex' in kwn:
            status = 'MAP_OK'
            target_page = '/unisex'
            target_type = 'Category page'
            why = 'Keyword generic danh mục unisex'
    elif intent_type == 'Brand / product commercial':
        brand_hit = None
        for b in brand_pages:
            if re.search(rf'(^| ){re.escape(b)}( |$)', kwn):
                brand_hit = b
                break
        if brand_hit:
            status = 'MAP_OK'
            target_page = brand_pages[brand_hit]
            target_type = 'Brand page'
            why = f'Khớp đúng brand page hiện có: {brand_hit}'
        elif missing_product:
            status = 'NEED_NEW_PRODUCT_PAGE'
            why = 'Là keyword product / commercial nhưng chưa có product page phù hợp trên web'
            note = f"Brand: {missing_product['brand']} | Candidate product: {missing_product['candidate']}"
    elif intent_type in ['Problem / need', 'Informational', 'Comparison', 'Product review', 'Trust', 'Commercial investigation']:
        if missing_product:
            status = 'NEED_NEW_PRODUCT_PAGE'
            why = 'Keyword thiên về sản phẩm cụ thể nhưng web chưa có page sản phẩm tương ứng'
            note = f"Brand: {missing_product['brand']} | Candidate product: {missing_product['candidate']}"
        else:
            status = 'NEED_NEW_CONTENT'
            why = 'Đúng intent SEO nhưng chưa có landing page / bài viết đúng chủ đề'
    else:
        status = 'UNMAPPED_REVIEW'
        why = 'Keyword mơ hồ, lỗi chính tả, nhiễu hoặc chưa đủ rõ để map an toàn'

    priority = priority_of(status, r['Volume'], r['CompetitionIndex'])
    rows.append({
        'Keyword': kw,
        'Volume': r['Volume'],
        'Competition': r['Competition'],
        'CompetitionIndex': r['CompetitionIndex'],
        'WordCount': r['WordCount'],
        'IntentType': intent_type,
        'KeywordCluster': cluster,
        'CurrentStatus': status,
        'TargetPage': target_page,
        'TargetPageType': target_type,
        'WhyMapped': why,
        'Note': note,
        'PriorityScore': priority,
    })

raw_df = pd.DataFrame(rows).sort_values(['CurrentStatus', 'PriorityScore', 'Volume', 'CompetitionIndex'], ascending=[True, True, False, True])
map_ok_df = raw_df[raw_df['CurrentStatus'] == 'MAP_OK'].copy().sort_values(['PriorityScore', 'Volume', 'CompetitionIndex'], ascending=[True, False, True])
new_content_df = raw_df[raw_df['CurrentStatus'] == 'NEED_NEW_CONTENT'].copy().sort_values(['Volume', 'CompetitionIndex', 'WordCount'], ascending=[False, True, False])
new_product_df = raw_df[raw_df['CurrentStatus'] == 'NEED_NEW_PRODUCT_PAGE'].copy().sort_values(['Volume', 'CompetitionIndex'], ascending=[False, True])
unmapped_df = raw_df[raw_df['CurrentStatus'] == 'UNMAPPED_REVIEW'].copy().sort_values(['Volume', 'CompetitionIndex'], ascending=[False, True])

# ---------- Cluster summary ----------
cluster_df = raw_df.groupby(['IntentType', 'KeywordCluster', 'CurrentStatus'], dropna=False).agg(
    KeywordCount=('Keyword', 'count'),
    TotalVolume=('Volume', 'sum')
).reset_index().sort_values(['CurrentStatus', 'TotalVolume'], ascending=[True, False])

# ---------- Daily content plan from NEED_NEW_CONTENT only ----------
content_candidates = new_content_df.copy()
content_candidates = content_candidates[(content_candidates['WordCount'] >= 3) & (content_candidates['Volume'] >= 10)]

picked = []
used_cluster = set()
for _, r in content_candidates.iterrows():
    cluster = r['KeywordCluster']
    if cluster in used_cluster and r['Volume'] < 100:
        continue
    used_cluster.add(cluster)
    title = ''
    slug = ''
    if cluster == 'Chính hãng / authentic / thật giả':
        title = 'Cách phân biệt nước hoa chính hãng, authentic và fake 2026'
        slug = 'phan-biet-nuoc-hoa-chinh-hang-authentic-fake-2026'
    elif cluster == 'EDT / EDP / Parfum':
        title = 'EDT, EDP, Parfum khác nhau gì? Cách chọn đúng trước khi mua 2026'
        slug = 'edt-edp-parfum-khac-nhau-gi-2026'
    elif cluster == 'Mini / chiết / dung tích nhỏ':
        title = 'Nước hoa mini, chiết 10ml, 20ml: nên mua loại nào 2026'
        slug = 'nuoc-hoa-mini-chiet-10ml-20ml-2026'
    elif cluster == 'Nhu cầu sử dụng':
        title = f"{r['Keyword'].capitalize()} — gợi ý chọn mùi phù hợp 2026"
        slug = norm(r['Keyword']).replace(' ', '-')[:70]
    elif cluster == 'Giải thích / how-to':
        title = f"{r['Keyword'].capitalize()} — giải thích nhanh và dễ hiểu 2026"
        slug = norm(r['Keyword']).replace(' ', '-')[:70]
    elif cluster == 'Review / đánh giá':
        title = f"{r['Keyword'].capitalize()} review 2026"
        slug = norm(r['Keyword']).replace(' ', '-')[:70]
    elif cluster == 'So sánh':
        title = f"{r['Keyword'].capitalize()} — so sánh nhanh 2026"
        slug = norm(r['Keyword']).replace(' ', '-')[:70]
    else:
        title = f"{r['Keyword'].capitalize()} 2026"
        slug = norm(r['Keyword']).replace(' ', '-')[:70]
    picked.append({
        'PrimaryKeyword': r['Keyword'],
        'Volume': r['Volume'],
        'Competition': r['Competition'],
        'CompetitionIndex': r['CompetitionIndex'],
        'IntentType': r['IntentType'],
        'KeywordCluster': r['KeywordCluster'],
        'SuggestedTitle': title,
        'SuggestedSlug': slug,
        'Reason': 'Keyword dài / đúng intent / chưa có landing page phù hợp',
    })
    if len(picked) >= 20:
        break

daily_plan_df = pd.DataFrame(picked)

# ---------- workbook ----------
with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    raw_df.to_excel(writer, sheet_name='01_Keyword_Raw_Cleaned', index=False)
    map_ok_df.to_excel(writer, sheet_name='02_MAP_OK_ExistingPages', index=False)
    new_content_df.to_excel(writer, sheet_name='03_NEED_NEW_CONTENT', index=False)
    new_product_df.to_excel(writer, sheet_name='04_NEED_NEW_PRODUCT_PAGE', index=False)
    unmapped_df.to_excel(writer, sheet_name='05_UNMAPPED_REVIEW', index=False)
    cluster_df.to_excel(writer, sheet_name='06_Content_Clusters', index=False)
    daily_plan_df.to_excel(writer, sheet_name='07_Daily_Content_Plan', index=False)

# ---------- format ----------
wb = load_workbook(out_path)
header_fill = PatternFill('solid', fgColor='1F4E78')
header_font = Font(color='FFFFFF', bold=True, name='Calibri')
body_font = Font(name='Calibri')
thin = Side(style='thin', color='D9E2F3')
for ws in wb.worksheets:
    ws.freeze_panes = 'A2'
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = body_font
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            val = '' if c.value is None else str(c.value)
            widths[c.column_letter] = min(max(widths.get(c.column_letter, 10), len(val) + 2), 42)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
wb.save(out_path)

summary = f"""# Keyword plan final\n\n- Tổng keyword: **{len(raw_df)}**\n- MAP_OK: **{len(map_ok_df)}**\n- NEED_NEW_CONTENT: **{len(new_content_df)}**\n- NEED_NEW_PRODUCT_PAGE: **{len(new_product_df)}**\n- UNMAPPED_REVIEW: **{len(unmapped_df)}**\n\n## Rule áp dụng\n- Không có page đúng intent -> không map bừa\n- Product keyword mà web chưa có sản phẩm -> đưa vào NEED_NEW_PRODUCT_PAGE\n- Keyword mơ hồ / lỗi / chưa chắc -> đưa vào UNMAPPED_REVIEW\n- Daily content plan chỉ lấy từ NEED_NEW_CONTENT\n"""
summary_path.write_text(summary, encoding='utf-8')
print(out_path)
print(summary_path)
